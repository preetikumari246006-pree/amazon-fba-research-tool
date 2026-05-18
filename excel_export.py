# excel_export.py — Handles all Excel output

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment
)
from openpyxl.utils import get_column_letter
import os
from config import OUTPUT_FOLDER, OUTPUT_FILE


def create_excel_report(all_products):
    """Creates a beautiful multi-sheet Excel report"""

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    wb = openpyxl.Workbook()

    # Create 3 sheets
    create_all_products_sheet(wb, all_products)
    create_top_opportunities_sheet(wb, all_products)
    create_summary_sheet(wb, all_products)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save file
    filepath = os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)
    wb.save(filepath)
    print(f"\n✅ Excel report saved: {filepath}")
    return filepath


# ─── SHEET 1: ALL PRODUCTS ────────────────────────────────

def create_all_products_sheet(wb, products):
    ws = wb.create_sheet("📦 All Products")

    headers = [
        "Keyword", "Title", "ASIN", "Price ($)",
        "Rating", "Reviews", "Best Seller",
        "Sponsored", "Profit Score", "URL"
    ]

    header_fill = PatternFill("solid", fgColor="2E86AB")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.get("Keyword"))
        ws.cell(row=row, column=2, value=product.get("Title"))
        ws.cell(row=row, column=3, value=product.get("ASIN"))
        ws.cell(row=row, column=4, value=product.get("Price ($)"))
        ws.cell(row=row, column=5, value=product.get("Rating"))
        ws.cell(row=row, column=6, value=product.get("Reviews"))
        ws.cell(row=row, column=7, value=product.get("Best Seller"))
        ws.cell(row=row, column=8, value=product.get("Sponsored"))
        ws.cell(row=row, column=9, value=product.get("Profit Score"))
        ws.cell(row=row, column=10, value=product.get("URL"))

        # Color rows by profit score
        score = product.get("Profit Score", 0)
        if score >= 70:
            row_fill = PatternFill("solid", fgColor="C8F7C5")  # Green
        elif score >= 40:
            row_fill = PatternFill("solid", fgColor="FFF9C4")  # Yellow
        else:
            row_fill = PatternFill("solid", fgColor="FFCDD2")  # Red

        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = row_fill

    auto_column_width(ws)


# ─── SHEET 2: TOP OPPORTUNITIES ───────────────────────────

def create_top_opportunities_sheet(wb, products):
    ws = wb.create_sheet("🏆 Top Opportunities")

    top = [
        p for p in products
        if (p.get("Profit Score", 0) >= 70
            and p.get("Sponsored") == "NO")
    ]

    top = sorted(top, key=lambda x: x.get("Profit Score", 0), reverse=True)

    headers = [
        "Rank", "Keyword", "Title", "Price ($)",
        "Rating", "Reviews", "Profit Score", "URL"
    ]

    header_fill = PatternFill("solid", fgColor="F4A261")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, product in enumerate(top, 2):
        ws.cell(row=row, column=1, value=row - 1)
        ws.cell(row=row, column=2, value=product.get("Keyword"))
        ws.cell(row=row, column=3, value=product.get("Title"))
        ws.cell(row=row, column=4, value=product.get("Price ($)"))
        ws.cell(row=row, column=5, value=product.get("Rating"))
        ws.cell(row=row, column=6, value=product.get("Reviews"))
        ws.cell(row=row, column=7, value=product.get("Profit Score"))
        ws.cell(row=row, column=8, value=product.get("URL"))

    auto_column_width(ws)

    if not top:
        ws.cell(row=2, column=1, value="No high-score products found yet.")


# ─── SHEET 3: SUMMARY ─────────────────────────────────────

def create_summary_sheet(wb, products):
    ws = wb.create_sheet("📊 Summary")

    total = len(products)
    avg_price = (
        sum(p["Price ($)"] for p in products if p.get("Price ($)"))
        / max(total, 1)
    )
    avg_rating = (
        sum(p["Rating"] for p in products if p.get("Rating"))
        / max(total, 1)
    )
    high_opportunity = sum(
        1 for p in products if p.get("Profit Score", 0) >= 70
    )
    best_sellers = sum(
        1 for p in products if p.get("Best Seller") == "YES"
    )

    title_cell = ws.cell(row=1, column=1, value="📊 Amazon FBA Research Summary")
    title_cell.font = Font(bold=True, size=14, color="2E86AB")

    summary_data = [
        ("Total Products Scraped", total),
        ("Average Price ($)", round(avg_price, 2)),
        ("Average Rating", round(avg_rating, 2)),
        ("High Opportunity Products (Score ≥ 70)", high_opportunity),
        ("Best Seller Products Found", best_sellers),
    ]

    label_font = Font(bold=True, size=11)
    for row, (label, value) in enumerate(summary_data, 3):
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)

    auto_column_width(ws)


# ─── HELPER ───────────────────────────────────────────────

def auto_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)
